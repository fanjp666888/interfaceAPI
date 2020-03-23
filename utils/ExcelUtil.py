#! /usr/bin/python3
# -*- coding: utf8 -*-
import datetime
import openpyxl
import logging
import os
import xlrd
import xlwt
from xlutils.copy import copy
from utils.Constant import Constant
from utils.LogSingleton import LogSingleton


logger = LogSingleton().getInstance()
logger = logging.getLogger(Constant.LOGGER_NAME)


class ExcelUtil(object):

    def __init__(self, Casepath):
        # print(os.path.join(os.getcwd(), Constant.PATH_FOR_FILES))
        self.excel_path = os.path.join(os.getcwd(), Constant.PATH_FOR_FILES)
        self.openExcel = self.load_excel(Casepath)
        self.sheetName = self.get_sheet_name()

    def load_excel(self, Casepath):
        """获取excel所有数据"""
        return openpyxl.load_workbook(self.excel_path + Casepath)  # "/Case/123456.xlsx"

    def get_sheet_name(self):
        """获取所有sheet名称"""
        return self.openExcel.sheetnames

    def get_sheet_data(self, index=None):
        """获取某一个sheet的内容"""
        if index is None:
            index = 0
        return self.openExcel[self.sheetName[index]]

    def get_cell_value(self, row, cols, index=None):
        """获取某个单元格的内容"""
        return self.get_sheet_data(index=index).cell(row, cols).value

    def get_rows(self, index=None):
        """获取总行数"""
        rows = 0
        row = self.get_sheet_data(index=index).max_row
        for i in range(row - 1):
            if self.get_cell_value(i + 1, 1):
                rows += 1
        return rows

    def get_rows_value(self, row, index=None):
        """获取某一行的内容"""
        return [i.value for i in self.get_sheet_data(index=index)[row]]

    def get_columns_value(self, key, index=None):
        """获取某一列的数据"""
        columns_list = []
        if key is None:
            key = 'A'
        columns_list_data = self.get_sheet_data(index=index)[key]
        for i in columns_list_data:
            columns_list.append(i.value)

        return columns_list

    def get_rows_number(self, case_id, key=None, index=None):
        """获取行号"""
        num = 1
        cols_data = self.get_columns_value(key, index=index)
        # print(cols_data, case_id)
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num += 1

        return num

        # ignoreTableHead = True
        # data = []
        # # workbook = xlrd.open_workbook(excelfile)
        # sheetData = self.get_sheet_data(index=None)
        # for i in range(sheetData.nrows):
        #     rowdata = []
        #     if ignoreTableHead == True and i == 0:
        #         continue
        #     for j in range(sheetData.ncols):
        #         cellValue = sheetData.cell(i, j).value
        #         if type(cellValue) == float:
        #             cellValue = str(int(cellValue))
        #         rowdata.append(cellValue)
        #     data.append(rowdata)
        # return data

    # def __init__(self):
    #     self.excel_path = os.path.join(os.getcwd(), Constant.PATH_FOR_REPORTS)
    #
    # # 获取excel文件sheet页里的特定列的内容
    # def getColumnData(self,excelfile,sheet,column):
    #     ignoreTableHead = True
    #     columnData = []
    #     workbook = xlrd.open_workbook(excelfile)
    #     sheetData = workbook.sheet_by_name(sheet)
    #     for i in range(sheetData.nrows):
    #         if ignoreTableHead == True and i == 0:
    #             continue
    #         cellValue = sheetData.cell(i,column).value
    #         if cellValue!='':columnData.append(sheetData.cell(i,column).value)
    #     return columnData
    #
    # # 获取excel文件sheet页里的特定行的内容
    # def getRowData(self,excelfile,sheet,row):
    #     rowData = []
    #     workbook = xlrd.open_workbook(excelfile)
    #     sheetData = workbook.sheet_by_name(sheet)
    #     for i in range(sheetData.ncols):
    #         rowData.append(sheetData.cell(row,i).value)
    #     return rowData
    #
    # 获取excel文件sheet页里的全部内容
    # def getAllData(self, excelfile, sheet):
    #     ignoreTableHead = True
    #     data = []
    #     workbook = xlrd.open_workbook(excelfile)
    #     sheetData = workbook.sheet_by_name(sheet)
    #     for i in range(sheetData.nrows):
    #         rowdata = []
    #         if ignoreTableHead == True and i == 0:
    #             continue
    #         for j in range(sheetData.ncols):
    #             cellValue = sheetData.cell(i, j).value
    #             if type(cellValue) == float:
    #                 cellValue = str(int(cellValue))
    #             rowdata.append(cellValue)
    #         data.append(rowdata)
    #     return data
    #
    # # 获取目录下的所有文件的绝对路径
    # def getAllfiles(self,foldername):
    #     excelfiles = []
    #     for files in os.walk(foldername):
    #         for file in files[2]:
    #             excelfiles.append(foldername+"\\"+file)
    #     return excelfiles
    #
    # # 仅获取目录下的所有excel文件的文件名
    # def getAllfilenames(self,foldername):
    #     allfiles = []
    #     for files in os.walk(foldername):
    #         for file in files[2]:
    #             allfiles.append(file)
    #     return allfiles
    #
    # # 把定制的结果写入到excel - jd
    # def exportToExcel(self,filename,searchresults):
    #     workbook = xlwt.Workbook(encoding='utf-8')
    #     sheet = workbook.add_sheet('Result')
    #     #sheet.write(0,1,'test')
    #     titles = ['搜索词','分词结果','品牌命中','相关分类','页数','召回总数','召回匹配']
    #     self.writeSheet(sheet,0,titles)
    #     search_id = 0
    #     for searchresult in searchresults:
    #         search_id += 1
    #         keyword = searchresult.getKeyword()
    #         words = searchresult.getWords()
    #         brand = searchresult.getBrand()
    #         category = searchresult.getCategory()
    #         totalcount = searchresult.getTotalCount()
    #         pagenum = searchresult.getPageNum()
    #         matchstatus = searchresult.getMatchStatus()
    #         self.writeSheet(sheet, search_id, [keyword,words,brand,category,pagenum,totalcount,matchstatus])
    #     workbook.save(self.excel_path+'\\'+filename+'.xls')
    # # 把定制结果写入到excel - suning_instock
    # def exportToExcel_instock(self,filename,searchresults):
    #     workbook = xlwt.Workbook(encoding='utf-8')
    #     sheet = workbook.add_sheet('Result')
    #     #sheet.write(0,1,'test')
    #     titles = ['搜索词','页数','召回总数','召回匹配','不匹配品']
    #     self.writeSheet(sheet,0,titles)
    #     search_id = 0
    #     for searchresult in searchresults:
    #         search_id += 1
    #         keyword = searchresult.getKeyword()
    #         totalcount = searchresult.getTotalCount()
    #         pagenum = searchresult.getPageNum()
    #         matchstatus = searchresult.getMatchStatus()
    #         unmatched = searchresult.getUnmatched()
    #         self.writeSheet(sheet, search_id, [keyword,pagenum,totalcount,matchstatus,unmatched])
    #     workbook.save(self.excel_path+'\\'+filename+'.xls')
    #
    # # 把定制的结果写入到excel - pptv
    # def exportVideoToExcel(self,filename,matchresult):
    #     workbook = xlwt.Workbook(encoding='utf-8')
    #     sheet = workbook.add_sheet('Result')
    #     #sheet.write(0,1,'test')
    #     titles = ['搜索词','完全匹配视频个数(前2页)','位置','分类','来源','视频Id']
    #     self.writeSheet(sheet,0,titles)
    #     search_id = 0
    #     for video in matchresult:
    #         search_id += 1
    #         keyword = video.getKeyword()
    #         videotype = ",".join(video.getType())
    #         count = video.getCount()
    #         position = ",".join(video.getPositions())
    #         source = ",".join(video.getSource())
    #         videoid = ",".join(video.getId())
    #         self.writeSheet(sheet, search_id, [keyword,count,position,videotype,source,videoid])
    #     workbook.save(self.excel_path+'\\'+filename+'.xls')
    #
    # # 把Diff结果写入到excel - pptv
    # def exportDiffResultToExcel(self,filename,list_changed,list_new,list_delete):
    #     workbook = xlwt.Workbook(encoding='utf-8')
    #     sheet = workbook.add_sheet('Result')
    #     titles = ['搜索词','正片个数','短视频个数','首个变动位置','对比结果','Deviation']
    #     self.writeSheet(sheet,0,titles)
    #     row_num = 0
    #     for video_c in list_changed:
    #         row_num += 1
    #         self.writeSheet(sheet, row_num, [video_c[0],video_c[1],video_c[2],video_c[3],"变动",video_c[4]])
    #
    #     for video_n in list_new:
    #         row_num += 1
    #         self.writeSheet(sheet, row_num, [video_n[0],video_n[1],video_n[2],video_n[3],"新增"])
    #
    #     for video_d in list_delete:
    #         row_num += 1
    #         self.writeSheet(sheet, row_num, [video_d[0],video_d[1],video_d[2],video_d[3],"删除"])
    #
    #     workbook.save(self.excel_path+'\\'+filename+'.xls')
    #
    # # 把结果写入到excel - pptv
    # def exportVideoSearchedResultToExcel(self,filename,keyword_result):
    #     workbook = xlwt.Workbook(encoding='utf-8')
    #     sheet = workbook.add_sheet('Result')
    #     sheet.write
    #     #startTime = (str(datetime.datetime.now())[:19]).replace(" ","-").replace(":","_")
    #     #sheet.write(0,1,'test')
    #     titles = ['搜索词','位置','标题','时间','简介','类型','地区','导演','主演','来源','视频Id']
    #     self.writeSheet(sheet,0,titles)
    #     row_num = 0
    #     for video in keyword_result:
    #         row_num += 1
    #         self.writeSheet(sheet, row_num,
    #                         [video.get_keyword(),
    #                          video.get_position(),
    #                          video.get_title(),
    #                          video.get_year(),
    #                          video.get_description(),
    #                          video.get_videotype(),
    #                          video.get_area(),
    #                          video.get_director(),
    #                          video.get_actor(),
    #                          video.get_source(),
    #                          video.get_videoid()])
    #         if not os.path.isdir(self.excel_path+'\\'+filename):
    #             os.mkdir(self.excel_path+'\\'+filename)
    #         workbook.save(self.excel_path+'\\'+filename+'\\'+self.replaceSpecialSymbol(video.get_keyword())+'.xls')
    #
    # # 把价格排序测试结果写入excel
    # def export_price_sort_test_result_to_excel(self, filename, test_result):
    #     workbook = xlrd.open_workbook(self.excel_path + '\\' + filename)
    #     sheets = workbook.sheet_names()
    #     worksheet = workbook.sheet_by_name(sheets[0])
    #     rows_old = worksheet.nrows
    #     new_workbook = copy(workbook)
    #     new_worksheet = new_workbook.get_sheet(0)
    #     self.excel_format(new_worksheet, 4)
    #     rows_new = 0
    #     for result in test_result:
    #         col_num = 0
    #         all_data = [rows_new + rows_old,
    #                     result.getKeyword(),
    #                     result.get_sort_type(),
    #                     result.get_sort_result(),
    #                     result.get_sort_result_detail().replace('<br>', '\n')]
    #         for data in all_data:
    #             new_worksheet.write(rows_new + rows_old, col_num, data)
    #             col_num += 1
    #         rows_new += 1
    #     new_workbook.save(os.path.abspath(self.excel_path + '\\' + filename))
    #
    # # 将商品类型判断结果写入excel
    # def export_commodity_type_judgment_results_to_excel(self, filename, test_result):
    #     workbook = xlrd.open_workbook(self.excel_path + '\\' + filename)
    #     sheets = workbook.sheet_names()
    #     worksheet = workbook.sheet_by_name(sheets[0])
    #     rows_old = worksheet.nrows
    #     new_workbook = copy(workbook)
    #     new_worksheet = new_workbook.get_sheet(0)
    #     self.excel_format(new_worksheet, 4)
    #     rows_new = 0
    #     for result in test_result:
    #         col_num = 0
    #         all_data = [rows_new + rows_old,
    #                     result.getKeyword(),
    #                     result.get_test_result(),
    #                     result.get_result_detail().replace('<br>', '\n')]
    #         for data in all_data:
    #             new_worksheet.write(rows_new + rows_old, col_num, data)
    #             col_num += 1
    #         rows_new += 1
    #     new_workbook.save(os.path.abspath(self.excel_path + '\\' + filename))
    #
    # # 设置excel单元格格式
    # def excel_format(self, worksheet, col_num):
    #     # 设置单元格宽度
    #     worksheet.col(col_num).width = 256*40
    #
    # # 写excel
    # def writeSheet(self,sheet,row_num,datalist):
    #     column = 0
    #     for data in datalist:
    #         sheet.write(row_num,column,data)
    #         column += 1
    # # 替换特殊符号,新建文件时不能含有这些特殊符号
    # def replaceSpecialSymbol(self,value):
    #     return value.replace("/","_").replace("|","_").replace("?","_").replace(":","_")